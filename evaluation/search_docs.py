import os
import glob
import pypdf
import docx
import openpyxl

DOCS_DIR = r"c:\Users\minhnn\Documents\cmcai\CAI_Legal\DocumentAndData"

def search_pdf(path):
    results = []
    try:
        reader = pypdf.PdfReader(path)
        for idx, page in enumerate(reader.pages, 1):
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                line_lower = line.lower()
                if any(x in line_lower for x in ["600", "database", "lưu", "mongo", "neo4j", "postgres", "sql", "server", "port", "host", "connection", "ip", "url"]):
                    results.append((idx, line.strip()))
    except Exception as e:
        print(f"Error reading PDF {path}: {e}")
    return results

def search_docx(path):
    results = []
    try:
        doc = docx.Document(path)
        for idx, para in enumerate(doc.paragraphs, 1):
            text = para.text
            if not text:
                continue
            line_lower = text.lower()
            if any(x in line_lower for x in ["600", "database", "lưu", "mongo", "neo4j", "postgres", "sql", "server", "port", "host", "connection", "ip", "url"]):
                results.append((idx, text.strip()))
    except Exception as e:
        print(f"Error reading DOCX {path}: {e}")
    return results

def search_xlsx(path):
    results = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_str = " ".join([str(cell) for cell in row if cell is not None])
                row_lower = row_str.lower()
                if any(x in row_lower for x in ["600", "database", "lưu", "mongo", "neo4j", "postgres", "sql", "server", "port", "host", "connection", "ip", "url"]):
                    results.append((sheet_name, row_idx, row_str[:200]))
    except Exception as e:
        print(f"Error reading XLSX {path}: {e}")
    return results

def main():
    import sys
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(f"Searching in {DOCS_DIR}...")
    files = glob.glob(os.path.join(DOCS_DIR, "*.*")) + glob.glob(os.path.join(DOCS_DIR, "Ref", "*.*"))
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        if ext == ".pdf":
            res = search_pdf(f)
            if res:
                print(f"\n=== FOUND IN PDF: {os.path.basename(f)} ({len(res)} matches) ===")
                for page, line in res[:15]:
                    print(f"  [Page {page}] {line}")
        elif ext == ".docx":
            res = search_docx(f)
            if res:
                print(f"\n=== FOUND IN DOCX: {os.path.basename(f)} ({len(res)} matches) ===")
                for para, line in res[:15]:
                    print(f"  [Para {para}] {line}")
        elif ext == ".xlsx":
            res = search_xlsx(f)
            if res:
                print(f"\n=== FOUND IN XLSX: {os.path.basename(f)} ({len(res)} matches) ===")
                for sheet, row, line in res[:15]:
                    print(f"  [{sheet} - Row {row}] {line}")

if __name__ == "__main__":
    main()
